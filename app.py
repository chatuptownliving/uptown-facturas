from flask import Flask, render_template, request, jsonify, send_file, redirect, session
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
from datetime import datetime
import anthropic
import json
import os
import io
import base64
import re
import xml.etree.ElementTree as ET

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'uptown-secret-2025')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///facturas.db').replace('postgres://', 'postgresql://')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

db = SQLAlchemy(app)
client = anthropic.Anthropic(api_key=os.environ.get('ANTHROPIC_API_KEY', ''))

ALLOWED_EXTENSIONS = {'xml', 'pdf', 'jpg', 'jpeg', 'png'}
MESES_NOMBRE = ['','Enero','Febrero','Marzo','Abril','Mayo','Junio',
                'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

MIME_MAP = {
    'pdf': 'application/pdf',
    'jpg': 'image/jpeg',
    'jpeg': 'image/jpeg',
    'png': 'image/png',
    'xml': 'application/xml'
}

# ─── MODELOS ────────────────────────────────────────────────────────────────

class Usuario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    facturas = db.relationship('Factura', backref='usuario', lazy=True)

class Factura(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    folio = db.Column(db.String(50))
    uuid_cfdi = db.Column(db.String(100))
    fecha_emision = db.Column(db.Date)
    mes = db.Column(db.Integer)
    anio = db.Column(db.Integer)
    rfc_emisor = db.Column(db.String(20))
    nombre_emisor = db.Column(db.String(200))
    rfc_receptor = db.Column(db.String(20))
    nombre_receptor = db.Column(db.String(200))
    concepto = db.Column(db.Text)
    subtotal = db.Column(db.Float, default=0)
    iva = db.Column(db.Float, default=0)
    total = db.Column(db.Float, default=0)
    moneda = db.Column(db.String(10), default='MXN')
    tipo = db.Column(db.String(20))
    estado_pago = db.Column(db.String(30), default='pendiente')
    archivo_nombre = db.Column(db.String(200))
    archivo_tipo = db.Column(db.String(10))
    archivo_contenido = db.Column(db.LargeBinary)
    drive_folder_id = db.Column(db.String(100))
    drive_file_id = db.Column(db.String(100))
    conciliada = db.Column(db.Boolean, default=False)
    movimiento_id = db.Column(db.Integer, db.ForeignKey('movimiento_bancario.id'), nullable=True)
    notas_ia = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class MovimientoBancario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    fecha = db.Column(db.Date)
    mes = db.Column(db.Integer)
    anio = db.Column(db.Integer)
    descripcion = db.Column(db.Text)
    referencia = db.Column(db.String(100))
    tipo = db.Column(db.String(10))
    monto = db.Column(db.Float)
    saldo = db.Column(db.Float, nullable=True)
    conciliado = db.Column(db.Boolean, default=False)
    factura_id = db.Column(db.Integer, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    facturas = db.relationship('Factura', backref='movimiento', foreign_keys=[Factura.movimiento_id])

# ─── HELPERS GENERALES ──────────────────────────────────────────────────────

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def limpiar_nombre(nombre):
    """Limpia un nombre para usarlo como nombre de archivo."""
    if not nombre:
        return 'Desconocido'
    nombre = nombre.strip()
    nombre = re.sub(r'[^\w\s\-]', '', nombre, flags=re.UNICODE)
    nombre = re.sub(r'\s+', '_', nombre)
    return nombre[:50]

def nombre_archivo_drive(factura, ext):
    """
    Genera el nombre del archivo para Drive:
    NombreEmpresa_YYYY-MM-DD.ext
    Para emitidas usa nombre_receptor, para recibidas usa nombre_emisor.
    """
    empresa = factura.nombre_emisor if factura.tipo == 'recibida' else factura.nombre_receptor
    empresa = limpiar_nombre(empresa) if empresa else 'Desconocido'
    fecha_str = factura.fecha_emision.strftime('%Y-%m-%d') if factura.fecha_emision else 'sin-fecha'
    return f"{empresa}_{fecha_str}.{ext}"

# ─── EXTRACCIÓN DE DATOS ────────────────────────────────────────────────────

def extraer_datos_xml(contenido_xml):
    try:
        root = ET.fromstring(contenido_xml)
        ns = {
            'cfdi': 'http://www.sat.gob.mx/cfd/4',
            'cfdi3': 'http://www.sat.gob.mx/cfd/3',
        }
        tag = root.tag
        prefix = 'cfdi' if '4' in tag else 'cfdi3'

        def attr(el, name, default=''):
            return el.get(name, default) if el is not None else default

        emisor = root.find(f'{prefix}:Emisor', ns) or root.find('Emisor')
        receptor = root.find(f'{prefix}:Receptor', ns) or root.find('Receptor')
        tfd = root.find('.//{http://www.sat.gob.mx/TimbreFiscalDigital}TimbreFiscalDigital')
        impuestos = root.find(f'{prefix}:Impuestos', ns) or root.find('Impuestos')

        fecha_str = attr(root, 'Fecha')
        fecha = datetime.strptime(fecha_str[:10], '%Y-%m-%d').date() if fecha_str else None

        iva = 0.0
        if impuestos is not None:
            traslados = impuestos.find(f'{prefix}:Traslados', ns) or impuestos.find('Traslados')
            if traslados is not None:
                for t in traslados:
                    iva += float(t.get('Importe', 0))

        # Extraer concepto del primer concepto
        concepto = ''
        conceptos = root.find(f'{prefix}:Conceptos', ns) or root.find('Conceptos')
        if conceptos is not None:
            primer = list(conceptos)[0] if len(list(conceptos)) > 0 else None
            if primer is not None:
                concepto = attr(primer, 'Descripcion') or attr(primer, 'descripcion', '')

        return {
            'folio': attr(root, 'Folio') or attr(root, 'Serie', '') + attr(root, 'Folio', ''),
            'uuid_cfdi': attr(tfd, 'UUID') if tfd is not None else '',
            'fecha_emision': fecha,
            'rfc_emisor': attr(emisor, 'Rfc') or attr(emisor, 'RFC'),
            'nombre_emisor': attr(emisor, 'Nombre'),
            'rfc_receptor': attr(receptor, 'Rfc') or attr(receptor, 'RFC'),
            'nombre_receptor': attr(receptor, 'Nombre'),
            'subtotal': float(attr(root, 'SubTotal', 0)),
            'iva': iva,
            'total': float(attr(root, 'Total', 0)),
            'moneda': attr(root, 'Moneda', 'MXN'),
            'concepto': concepto,
            'fuente': 'xml'
        }
    except Exception as e:
        return {'error': str(e)}

def extraer_datos_ia(contenido_bytes, tipo_archivo, nombre_archivo):
    try:
        if tipo_archivo in ['jpg', 'jpeg', 'png']:
            media_type = MIME_MAP.get(tipo_archivo, 'image/jpeg')
            b64 = base64.standard_b64encode(contenido_bytes).decode('utf-8')
            content = [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}},
                {"type": "text", "text": """Analiza esta imagen de factura mexicana y extrae ÚNICAMENTE un JSON con estos campos (sin texto adicional, sin ```):
{"folio":"","uuid_cfdi":"","fecha_emision":"YYYY-MM-DD","rfc_emisor":"","nombre_emisor":"","rfc_receptor":"","nombre_receptor":"","concepto":"","subtotal":0.0,"iva":0.0,"total":0.0,"moneda":"MXN"}
Si no encuentras un campo déjalo vacío o en 0. Fecha siempre YYYY-MM-DD."""}
            ]
        else:
            b64 = base64.standard_b64encode(contenido_bytes).decode('utf-8')
            content = [
                {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}},
                {"type": "text", "text": """Analiza este PDF de factura mexicana y extrae ÚNICAMENTE un JSON con estos campos (sin texto adicional, sin ```):
{"folio":"","uuid_cfdi":"","fecha_emision":"YYYY-MM-DD","rfc_emisor":"","nombre_emisor":"","rfc_receptor":"","nombre_receptor":"","concepto":"","subtotal":0.0,"iva":0.0,"total":0.0,"moneda":"MXN"}
Si no encuentras un campo déjalo vacío o en 0. Fecha siempre YYYY-MM-DD."""}
            ]

        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1000,
            messages=[{"role": "user", "content": content}]
        )

        texto = response.content[0].text.strip().replace('```json', '').replace('```', '').strip()
        datos = json.loads(texto)

        if datos.get('fecha_emision'):
            try:
                datos['fecha_emision'] = datetime.strptime(datos['fecha_emision'][:10], '%Y-%m-%d').date()
            except:
                datos['fecha_emision'] = None

        datos['fuente'] = 'ia'
        return datos
    except Exception as e:
        return {'error': str(e)}

# ─── GOOGLE DRIVE OAUTH ─────────────────────────────────────────────────────

def get_drive_service():
    """Retorna servicio Drive autenticado via OAuth token guardado en session."""
    try:
        from google.oauth2.credentials import Credentials
        from googleapiclient.discovery import build
        token_data = session.get('drive_token')
        if not token_data:
            return None
        creds = Credentials(
            token=token_data.get('access_token'),
            refresh_token=token_data.get('refresh_token'),
            token_uri='https://oauth2.googleapis.com/token',
            client_id=os.environ.get('GOOGLE_CLIENT_ID'),
            client_secret=os.environ.get('GOOGLE_CLIENT_SECRET'),
            scopes=['https://www.googleapis.com/auth/drive']
        )
        return build('drive', 'v3', credentials=creds)
    except Exception:
        return None

def buscar_o_crear_carpeta(service, nombre, parent_id=None):
    """Busca carpeta por nombre, la crea si no existe."""
    q = f"name='{nombre}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    if parent_id:
        q += f" and '{parent_id}' in parents"
    res = service.files().list(q=q, fields='files(id,name)').execute()
    files = res.get('files', [])
    if files:
        return files[0]['id']
    meta = {'name': nombre, 'mimeType': 'application/vnd.google-apps.folder'}
    if parent_id:
        meta['parents'] = [parent_id]
    f = service.files().create(body=meta, fields='id').execute()
    return f.get('id')

def subir_archivo_drive(service, contenido_bytes, nombre_archivo, mime_type, folder_id):
    """Sube un archivo a Drive dentro de folder_id."""
    from googleapiclient.http import MediaIoBaseUpload
    meta = {'name': nombre_archivo, 'parents': [folder_id]}
    media = MediaIoBaseUpload(io.BytesIO(contenido_bytes), mimetype=mime_type)
    f = service.files().create(body=meta, media_body=media, fields='id,webViewLink').execute()
    return f.get('id'), f.get('webViewLink', '')

def organizar_factura_drive(service, factura, contenido_bytes, ext):
    """
    Sube factura a Drive con estructura:
    Facturas / YYYY / Mes / NombreEmpresa_YYYY-MM-DD.ext
    """
    try:
        # Carpeta raíz: Facturas
        root_id = os.environ.get('DRIVE_FOLDER_ID')
        carpeta_facturas = buscar_o_crear_carpeta(service, 'Facturas', root_id)

        # Carpeta año
        anio_str = str(factura.anio) if factura.anio else 'Sin_Fecha'
        carpeta_anio = buscar_o_crear_carpeta(service, anio_str, carpeta_facturas)

        # Carpeta mes
        mes_str = MESES_NOMBRE[factura.mes] if factura.mes else 'Sin_Mes'
        carpeta_mes = buscar_o_crear_carpeta(service, mes_str, carpeta_anio)

        # Nombre del archivo: Empresa_Fecha.ext
        nombre = nombre_archivo_drive(factura, ext)
        mime = MIME_MAP.get(ext, 'application/octet-stream')

        file_id, link = subir_archivo_drive(service, contenido_bytes, nombre, mime, carpeta_mes)

        factura.drive_folder_id = carpeta_mes
        factura.drive_file_id = file_id
        db.session.commit()

        return {'ok': True, 'file_id': file_id, 'nombre': nombre, 'link': link}
    except Exception as e:
        return {'ok': False, 'error': str(e)}

# ─── RUTAS OAUTH ────────────────────────────────────────────────────────────

@app.route('/oauth/login')
def oauth_login():
    from urllib.parse import urlencode
    params = {
        'client_id': os.environ.get('GOOGLE_CLIENT_ID'),
        'redirect_uri': os.environ.get('OAUTH_REDIRECT', 'https://facturas.uptownliving.mx/oauth/callback'),
        'response_type': 'code',
        'scope': 'https://www.googleapis.com/auth/drive',
        'access_type': 'offline',
        'prompt': 'consent'
    }
    url = 'https://accounts.google.com/o/oauth2/v2/auth?' + urlencode(params)
    return redirect(url)

@app.route('/oauth/callback')
def oauth_callback():
    import urllib.request
    import urllib.parse
    code = request.args.get('code')
    if not code:
        return '<h2>Error: no se recibió código de autorización</h2>', 400

    data = urllib.parse.urlencode({
        'code': code,
        'client_id': os.environ.get('GOOGLE_CLIENT_ID'),
        'client_secret': os.environ.get('GOOGLE_CLIENT_SECRET'),
        'redirect_uri': os.environ.get('OAUTH_REDIRECT', 'https://facturas.uptownliving.mx/oauth/callback'),
        'grant_type': 'authorization_code'
    }).encode()

    req = urllib.request.Request('https://oauth2.googleapis.com/token', data=data)
    try:
        with urllib.request.urlopen(req) as resp:
            token_data = json.loads(resp.read().decode())
        session['drive_token'] = token_data
        return redirect('/?drive=ok')
    except Exception as e:
        return f'<h2>Error obteniendo token: {e}</h2>', 500

@app.route('/oauth/logout')
def oauth_logout():
    session.pop('drive_token', None)
    return redirect('/')

@app.route('/api/drive/status', methods=['GET'])
def drive_status():
    service = get_drive_service()
    return jsonify({'conectado': service is not None})

@app.route('/api/drive/organizar-todas', methods=['POST'])
def organizar_todas_drive():
    """Organiza en Drive todas las facturas que tienen contenido guardado."""
    service = get_drive_service()
    if not service:
        return jsonify({'error': 'Drive no autorizado. Ve a /oauth/login'}), 401

    usuario_id = request.json.get('usuario_id', 1)
    facturas = Factura.query.filter_by(usuario_id=usuario_id).filter(
        Factura.archivo_contenido.isnot(None),
        Factura.drive_file_id.is_(None)
    ).all()

    subidas = 0
    errores = 0
    for f in facturas:
        ext = f.archivo_tipo or 'pdf'
        res = organizar_factura_drive(service, f, f.archivo_contenido, ext)
        if res['ok']:
            subidas += 1
        else:
            errores += 1

    return jsonify({'subidas': subidas, 'errores': errores})

# ─── RUTAS PRINCIPALES ──────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/usuarios', methods=['GET'])
def get_usuarios():
    usuarios = Usuario.query.all()
    return jsonify([{'id': u.id, 'nombre': u.nombre, 'email': u.email} for u in usuarios])

@app.route('/api/usuarios', methods=['POST'])
def crear_usuario():
    data = request.json
    if Usuario.query.filter_by(email=data['email']).first():
        return jsonify({'error': 'Email ya registrado'}), 400
    u = Usuario(nombre=data['nombre'], email=data['email'])
    db.session.add(u)
    db.session.commit()
    return jsonify({'id': u.id, 'nombre': u.nombre, 'email': u.email}), 201

# ─── RUTAS FACTURAS ─────────────────────────────────────────────────────────

@app.route('/api/facturas', methods=['GET'])
def get_facturas():
    usuario_id = request.args.get('usuario_id', 1, type=int)
    q = request.args.get('q', '')
    tipo = request.args.get('tipo', '')
    mes = request.args.get('mes', type=int)
    anio = request.args.get('anio', type=int)
    estado = request.args.get('estado', '')
    conciliada = request.args.get('conciliada')

    query = Factura.query.filter_by(usuario_id=usuario_id)

    if q:
        like = f'%{q}%'
        query = query.filter(db.or_(
            Factura.folio.ilike(like),
            Factura.nombre_emisor.ilike(like),
            Factura.nombre_receptor.ilike(like),
            Factura.rfc_emisor.ilike(like),
            Factura.rfc_receptor.ilike(like),
            Factura.concepto.ilike(like)
        ))
    if tipo: query = query.filter_by(tipo=tipo)
    if mes: query = query.filter_by(mes=mes)
    if anio: query = query.filter_by(anio=anio)
    if estado: query = query.filter_by(estado_pago=estado)
    if conciliada is not None:
        query = query.filter_by(conciliada=(conciliada == 'true'))

    facturas = query.order_by(Factura.fecha_emision.desc()).all()

    return jsonify([{
        'id': f.id, 'folio': f.folio, 'uuid_cfdi': f.uuid_cfdi,
        'fecha_emision': f.fecha_emision.isoformat() if f.fecha_emision else None,
        'mes': f.mes, 'anio': f.anio,
        'rfc_emisor': f.rfc_emisor, 'nombre_emisor': f.nombre_emisor,
        'rfc_receptor': f.rfc_receptor, 'nombre_receptor': f.nombre_receptor,
        'concepto': f.concepto, 'subtotal': f.subtotal, 'iva': f.iva,
        'total': f.total, 'moneda': f.moneda, 'tipo': f.tipo,
        'estado_pago': f.estado_pago, 'conciliada': f.conciliada,
        'archivo_nombre': f.archivo_nombre, 'archivo_tipo': f.archivo_tipo,
        'drive_file_id': f.drive_file_id, 'notas_ia': f.notas_ia,
        'created_at': f.created_at.isoformat()
    } for f in facturas])

@app.route('/api/facturas', methods=['POST'])
def subir_factura():
    usuario_id = request.form.get('usuario_id', 1, type=int)
    tipo = request.form.get('tipo', 'emitida')

    if 'archivo' not in request.files:
        return jsonify({'error': 'No se envió archivo'}), 400

    archivo = request.files['archivo']
    if not archivo.filename or not allowed_file(archivo.filename):
        return jsonify({'error': 'Archivo no válido'}), 400

    filename = secure_filename(archivo.filename)
    ext = filename.rsplit('.', 1)[1].lower()
    contenido = archivo.read()

    if ext == 'xml':
        datos = extraer_datos_xml(contenido.decode('utf-8', errors='ignore'))
    else:
        datos = extraer_datos_ia(contenido, ext, filename)

    if 'error' in datos:
        return jsonify({'error': f'Error extrayendo datos: {datos["error"]}'}), 422

    fecha = datos.get('fecha_emision')

    factura = Factura(
        usuario_id=usuario_id, tipo=tipo,
        folio=datos.get('folio', ''), uuid_cfdi=datos.get('uuid_cfdi', ''),
        fecha_emision=fecha,
        mes=fecha.month if fecha else None,
        anio=fecha.year if fecha else None,
        rfc_emisor=datos.get('rfc_emisor', ''), nombre_emisor=datos.get('nombre_emisor', ''),
        rfc_receptor=datos.get('rfc_receptor', ''), nombre_receptor=datos.get('nombre_receptor', ''),
        concepto=datos.get('concepto', ''),
        subtotal=datos.get('subtotal', 0), iva=datos.get('iva', 0), total=datos.get('total', 0),
        moneda=datos.get('moneda', 'MXN'),
        archivo_nombre=filename, archivo_tipo=ext,
        archivo_contenido=contenido,
        notas_ia=f'Extraído por {"XML parser" if ext == "xml" else "IA (Claude)"}'
    )

    db.session.add(factura)
    db.session.commit()

    # Subir a Drive automáticamente si está autorizado
    service = get_drive_service()
    drive_resultado = None
    if service:
        drive_resultado = organizar_factura_drive(service, factura, contenido, ext)

    return jsonify({
        'id': factura.id, 'folio': factura.folio, 'total': factura.total,
        'fecha_emision': factura.fecha_emision.isoformat() if factura.fecha_emision else None,
        'nombre_emisor': factura.nombre_emisor, 'nombre_receptor': factura.nombre_receptor,
        'tipo': factura.tipo, 'fuente_extraccion': datos.get('fuente', 'desconocida'),
        'drive': drive_resultado
    }), 201

@app.route('/api/facturas/manual', methods=['POST'])
def agregar_factura_manual():
    data = request.json
    fecha = None
    if data.get('fecha_emision'):
        try:
            fecha = datetime.strptime(data['fecha_emision'], '%Y-%m-%d').date()
        except: pass

    factura = Factura(
        usuario_id=data.get('usuario_id', 1),
        folio=data.get('folio', ''), fecha_emision=fecha,
        mes=fecha.month if fecha else None, anio=fecha.year if fecha else None,
        rfc_emisor=data.get('rfc_emisor', ''), nombre_emisor=data.get('nombre_emisor', ''),
        rfc_receptor=data.get('rfc_receptor', ''), nombre_receptor=data.get('nombre_receptor', ''),
        concepto=data.get('concepto', ''),
        subtotal=float(data.get('subtotal', 0)), iva=float(data.get('iva', 0)),
        total=float(data.get('total', 0)), moneda=data.get('moneda', 'MXN'),
        tipo=data.get('tipo', 'emitida'), archivo_tipo='manual', notas_ia='Ingreso manual'
    )
    db.session.add(factura)
    db.session.commit()
    return jsonify({'id': factura.id, 'folio': factura.folio}), 201

@app.route('/api/facturas/<int:factura_id>', methods=['DELETE'])
def eliminar_factura(factura_id):
    f = Factura.query.get_or_404(factura_id)
    db.session.delete(f)
    db.session.commit()
    return jsonify({'ok': True})

# ─── RUTAS MOVIMIENTOS ──────────────────────────────────────────────────────

@app.route('/api/movimientos', methods=['GET'])
def get_movimientos():
    usuario_id = request.args.get('usuario_id', 1, type=int)
    mes = request.args.get('mes', type=int)
    anio = request.args.get('anio', type=int)
    query = MovimientoBancario.query.filter_by(usuario_id=usuario_id)
    if mes: query = query.filter_by(mes=mes)
    if anio: query = query.filter_by(anio=anio)
    movs = query.order_by(MovimientoBancario.fecha.desc()).all()
    return jsonify([{
        'id': m.id, 'fecha': m.fecha.isoformat() if m.fecha else None,
        'descripcion': m.descripcion, 'referencia': m.referencia,
        'tipo': m.tipo, 'monto': m.monto, 'saldo': m.saldo,
        'conciliado': m.conciliado, 'factura_id': m.factura_id
    } for m in movs])

@app.route('/api/movimientos', methods=['POST'])
def agregar_movimiento():
    data = request.json
    fecha = None
    if data.get('fecha'):
        try: fecha = datetime.strptime(data['fecha'], '%Y-%m-%d').date()
        except: pass
    m = MovimientoBancario(
        usuario_id=data.get('usuario_id', 1), fecha=fecha,
        mes=fecha.month if fecha else None, anio=fecha.year if fecha else None,
        descripcion=data.get('descripcion', ''), referencia=data.get('referencia', ''),
        tipo=data.get('tipo', 'abono'), monto=float(data.get('monto', 0)),
        saldo=float(data.get('saldo', 0)) if data.get('saldo') else None
    )
    db.session.add(m)
    db.session.commit()
    return jsonify({'id': m.id}), 201

@app.route('/api/movimientos/importar-csv', methods=['POST'])
def importar_csv():
    usuario_id = request.form.get('usuario_id', 1, type=int)
    if 'archivo' not in request.files:
        return jsonify({'error': 'No se envió archivo'}), 400
    archivo = request.files['archivo']
    contenido = archivo.read().decode('utf-8-sig', errors='ignore')
    lineas = contenido.strip().split('\n')
    importados = 0
    for linea in lineas[1:]:
        partes = linea.strip().split(',')
        if len(partes) < 3: continue
        try:
            fecha_str = partes[0].strip().strip('"')
            desc = partes[1].strip().strip('"')
            monto_raw = float(partes[2].strip().strip('"').replace('$', '').replace(',', ''))
            tipo = 'abono' if monto_raw >= 0 else 'cargo'
            monto = abs(monto_raw)
            saldo = float(partes[3].strip().strip('"').replace('$', '').replace(',', '')) if len(partes) > 3 else None
            fecha = datetime.strptime(fecha_str[:10], '%Y-%m-%d').date()
            m = MovimientoBancario(
                usuario_id=usuario_id, fecha=fecha,
                mes=fecha.month, anio=fecha.year,
                descripcion=desc, tipo=tipo, monto=monto, saldo=saldo
            )
            db.session.add(m)
            importados += 1
        except: continue
    db.session.commit()
    return jsonify({'importados': importados})


@app.route('/api/estados-cuenta/subir', methods=['POST'])
def subir_estado_cuenta():
    usuario_id = request.form.get('usuario_id', 1, type=int)
    mes = request.form.get('mes', type=int)
    anio = request.form.get('anio', type=int)
    banco = request.form.get('banco', 'Banco')

    if 'archivo' not in request.files:
        return jsonify({"error": "No se envio archivo"}), 400

    archivo = request.files['archivo']
    if not archivo.filename:
        return jsonify({"error": "Archivo vacio"}), 400

    filename = secure_filename(archivo.filename)
    ext = filename.rsplit(".", 1)[1].lower() if "." in filename else "pdf"
    contenido = archivo.read()

    # Subir a Drive si está autorizado
    drive_resultado = None
    service = get_drive_service()
    if service:
        try:
            root_id = os.environ.get("DRIVE_FOLDER_ID")
            carpeta_edos = buscar_o_crear_carpeta(service, "Estados de Cuenta", root_id)
            anio_str = str(anio) if anio else "Sin_Anio"
            carpeta_anio = buscar_o_crear_carpeta(service, anio_str, carpeta_edos)
            mes_str = MESES_NOMBRE[mes] if mes and 1 <= mes <= 12 else "Sin_Mes"
            carpeta_mes = buscar_o_crear_carpeta(service, mes_str, carpeta_anio)
            nombre_archivo = f"{limpiar_nombre(banco)}_{anio_str}-{str(mes).zfill(2) if mes else '00'}.{ext}"
            mime = MIME_MAP.get(ext, "application/octet-stream")
            file_id, link = subir_archivo_drive(service, contenido, nombre_archivo, mime, carpeta_mes)
            drive_resultado = {"ok": True, "nombre": nombre_archivo, "link": link}
        except Exception as e:
            drive_resultado = {"ok": False, "error": str(e)}

    return jsonify({
        "ok": True,
        "filename": filename,
        "drive": drive_resultado
    }), 201

# ─── CONCILIACIÓN ───────────────────────────────────────────────────────────

def conciliar_automatico(usuario_id):
    facturas = Factura.query.filter_by(usuario_id=usuario_id, conciliada=False).all()
    movimientos = MovimientoBancario.query.filter_by(usuario_id=usuario_id, conciliado=False).all()
    conciliados = 0
    for f in facturas:
        tipo_mov = 'abono' if f.tipo == 'emitida' else 'cargo'
        mejor_match = None
        mejor_diff = float('inf')
        for m in movimientos:
            if m.tipo != tipo_mov: continue
            if f.fecha_emision and m.fecha:
                if abs((f.fecha_emision - m.fecha).days) > 60: continue
            diff = abs(m.monto - f.total)
            if diff <= f.total * 0.02 and diff < mejor_diff:
                mejor_diff = diff
                mejor_match = m
        if mejor_match:
            f.conciliada = True
            f.movimiento_id = mejor_match.id
            f.estado_pago = 'pagada'
            mejor_match.conciliado = True
            mejor_match.factura_id = f.id
            movimientos.remove(mejor_match)
            conciliados += 1
    db.session.commit()
    return conciliados

@app.route('/api/conciliar', methods=['POST'])
def conciliar():
    usuario_id = request.json.get('usuario_id', 1)
    n = conciliar_automatico(usuario_id)
    return jsonify({'conciliados': n})

@app.route('/api/conciliacion/resumen', methods=['GET'])
def resumen_conciliacion():
    usuario_id = request.args.get('usuario_id', 1, type=int)
    mes = request.args.get('mes', type=int)
    anio = request.args.get('anio', type=int)
    q = Factura.query.filter_by(usuario_id=usuario_id)
    if mes: q = q.filter_by(mes=mes)
    if anio: q = q.filter_by(anio=anio)
    facturas = q.all()
    emitidas = [f for f in facturas if f.tipo == 'emitida']
    recibidas = [f for f in facturas if f.tipo == 'recibida']
    total_emitido = sum(f.total for f in emitidas)
    total_recibido = sum(f.total for f in recibidas)
    return jsonify({
        'total_facturas': len(facturas),
        'emitidas': len(emitidas), 'recibidas': len(recibidas),
        'conciliadas': sum(1 for f in facturas if f.conciliada),
        'pendientes': sum(1 for f in facturas if not f.conciliada),
        'total_emitido': total_emitido, 'total_recibido': total_recibido,
        'total_cobrado': sum(f.total for f in emitidas if f.conciliada),
        'total_pagado': sum(f.total for f in recibidas if f.conciliada),
        'por_cobrar': sum(f.total for f in emitidas if not f.conciliada),
        'por_pagar': sum(f.total for f in recibidas if not f.conciliada),
        'balance': total_emitido - total_recibido
    })

# ─── EXPORTAR EXCEL ─────────────────────────────────────────────────────────

@app.route('/api/exportar/excel', methods=['GET'])
def exportar_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'error': 'openpyxl no instalado'}), 500

    usuario_id = request.args.get('usuario_id', 1, type=int)
    mes = request.args.get('mes', type=int)
    anio = request.args.get('anio', type=int)

    q = Factura.query.filter_by(usuario_id=usuario_id)
    if mes: q = q.filter_by(mes=mes)
    if anio: q = q.filter_by(anio=anio)
    facturas = q.order_by(Factura.fecha_emision).all()

    qm = MovimientoBancario.query.filter_by(usuario_id=usuario_id)
    if mes: qm = qm.filter_by(mes=mes)
    if anio: qm = qm.filter_by(anio=anio)
    movimientos = qm.order_by(MovimientoBancario.fecha).all()

    wb = openpyxl.Workbook()
    hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    center = Alignment(horizontal='center', vertical='center')
    thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    fill_ok = PatternFill('solid', start_color='FFF3E0', end_color='FFF3E0')
    fill_warn = PatternFill('solid', start_color='FAEEDA', end_color='FAEEDA')
    fill_danger = PatternFill('solid', start_color='FCEBEB', end_color='FCEBEB')
    orange_fill = PatternFill('solid', start_color='E18F00', end_color='E18F00')
    dark_fill = PatternFill('solid', start_color='393939', end_color='393939')

    def set_header(ws, headers, fill):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = hdr_font
            cell.fill = fill
            cell.alignment = center
            cell.border = thin

    emitidas = [f for f in facturas if f.tipo == 'emitida']
    recibidas = [f for f in facturas if f.tipo == 'recibida']

    # ── Hoja 1: Resumen ──
    ws_res = wb.active
    ws_res.title = 'Resumen'
    ws_res['A1'] = 'UPTOWN — REPORTE DE CONCILIACIÓN'
    ws_res['A1'].font = Font(name='Arial', bold=True, size=14, color='E18F00')
    ws_res['A2'] = f'Período: {MESES_NOMBRE[mes] if mes else "Todo"} {anio or ""}'
    ws_res['A2'].font = Font(name='Arial', size=11, color='393939')
    ws_res['A3'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws_res['A3'].font = Font(name='Arial', size=10, color='888780')
    ws_res.append(['', '', ''])
    ws_res.append(['CONCEPTO', 'CANTIDAD', 'MONTO'])
    for cell in ws_res[5]:
        cell.font = hdr_font
        cell.fill = orange_fill
        cell.alignment = center
    datos = [
        ['Total facturas emitidas', len(emitidas), sum(f.total for f in emitidas)],
        ['Total facturas recibidas', len(recibidas), sum(f.total for f in recibidas)],
        ['Facturas conciliadas', sum(1 for f in facturas if f.conciliada), sum(f.total for f in facturas if f.conciliada)],
        ['Por cobrar', sum(1 for f in emitidas if not f.conciliada), sum(f.total for f in emitidas if not f.conciliada)],
        ['Por pagar', sum(1 for f in recibidas if not f.conciliada), sum(f.total for f in recibidas if not f.conciliada)],
        ['BALANCE NETO', '', sum(f.total for f in emitidas) - sum(f.total for f in recibidas)],
    ]
    for row in datos:
        ws_res.append(row)
    for row in ws_res.iter_rows(min_row=6, max_row=11):
        for cell in row:
            cell.border = thin
            if cell.column == 3 and isinstance(cell.value, (int, float)):
                cell.number_format = '"$"#,##0.00'
    for cell in ws_res[11]:
        cell.font = Font(name='Arial', bold=True, color='FFFFFF')
        cell.fill = dark_fill
    ws_res.column_dimensions['A'].width = 32
    ws_res.column_dimensions['B'].width = 14
    ws_res.column_dimensions['C'].width = 18

    # ── Hoja 2: Facturas ──
    ws_fac = wb.create_sheet('Facturas')
    set_header(ws_fac, ['Folio','Fecha','Tipo','Emisor','RFC Emisor',
        'Receptor','RFC Receptor','Concepto','Subtotal','IVA','Total','Moneda','Estado','Conciliada','En Drive'], orange_fill)
    for f in facturas:
        ws_fac.append([
            f.folio, f.fecha_emision, f.tipo,
            f.nombre_emisor, f.rfc_emisor,
            f.nombre_receptor, f.rfc_receptor, f.concepto,
            f.subtotal, f.iva, f.total, f.moneda,
            f.estado_pago, 'Sí' if f.conciliada else 'No',
            'Sí' if f.drive_file_id else 'No'
        ])
        last = ws_fac.max_row
        fill = fill_ok if f.conciliada else (fill_warn if f.estado_pago == 'pendiente' else fill_danger)
        for cell in ws_fac[last]:
            cell.border = thin
            cell.fill = fill
        for col in [9, 10, 11]:
            ws_fac.cell(last, col).number_format = '"$"#,##0.00'
    for i, w in enumerate([10,12,10,30,15,30,15,35,14,14,14,8,12,10,10], 1):
        ws_fac.column_dimensions[get_column_letter(i)].width = w

    # ── Hoja 3: Estado de cuenta ──
    ws_edo = wb.create_sheet('Estado de Cuenta')
    set_header(ws_edo, ['Fecha','Descripción','Referencia','Tipo','Monto','Saldo','Conciliado','Factura ID'], dark_fill)
    for m in movimientos:
        ws_edo.append([m.fecha, m.descripcion, m.referencia, m.tipo,
            m.monto, m.saldo, 'Sí' if m.conciliado else 'No', m.factura_id or ''])
        last = ws_edo.max_row
        for cell in ws_edo[last]:
            cell.border = thin
            cell.fill = fill_ok if m.conciliado else fill_warn
        ws_edo.cell(last, 5).number_format = '"$"#,##0.00'
        if m.saldo: ws_edo.cell(last, 6).number_format = '"$"#,##0.00'
    for i, w in enumerate([12,40,20,10,14,14,12,12], 1):
        ws_edo.column_dimensions[get_column_letter(i)].width = w

    # ── Hoja 4: Conciliación ──
    ws_con = wb.create_sheet('Conciliación')
    set_header(ws_con, ['Folio','Fecha','Empresa','Tipo','Total Factura','Monto Banco','Diferencia','Estado'], orange_fill)
    for f in facturas:
        mov = f.movimiento
        monto_banco = mov.monto if mov else 0
        diff = monto_banco - f.total if mov else -f.total
        estado = 'Conciliada' if f.conciliada else ('Sin cobro' if f.tipo == 'emitida' else 'Sin pago')
        empresa = f.nombre_emisor if f.tipo == 'recibida' else f.nombre_receptor
        ws_con.append([f.folio, f.fecha_emision, empresa, f.tipo,
            f.total, monto_banco or '', diff if mov else '', estado])
        last = ws_con.max_row
        fill = fill_ok if f.conciliada else (fill_warn if f.tipo == 'recibida' else fill_danger)
        for cell in ws_con[last]:
            cell.border = thin
            cell.fill = fill
        for col in [5, 6, 7]:
            if ws_con.cell(last, col).value != '':
                ws_con.cell(last, col).number_format = '"$"#,##0.00'
    for i, w in enumerate([14,14,32,10,16,16,14,14], 1):
        ws_con.column_dimensions[get_column_letter(i)].width = w

    for ws in [ws_fac, ws_edo, ws_con]:
        ws.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    nombre = f'uptown_conciliacion_{anio or "all"}_{mes or "all"}.xlsx'
    return send_file(output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=nombre)

# ─── INIT ────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        if not Usuario.query.first():
            db.session.add(Usuario(nombre='Admin', email='admin@empresa.com'))
            db.session.commit()
    app.run(debug=True, port=5001)
